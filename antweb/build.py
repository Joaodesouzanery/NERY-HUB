#!/usr/bin/env python3
"""
build.py — Pipeline AntWeb + Serrapilheira (Observação Focal)

O que faz
---------
1. Lê a planilha de observação focal de formigas (Serrapilheira) e a ORGANIZA
   (normaliza subfamília/gênero/espécie, corrige separadores, remove linhas 'na',
   padroniza nomes de área, tipa horas/números).
2. Extrai os táxons presentes (subfamília > gênero > morfoespécie).
3. Consulta a API do AntWeb para cada táxon e puxa TODOS os specimens
   disponíveis (nomes, taxonomia, localidade, metadados e URLs de imagens),
   incluindo a "Favorite Image" (imagem representativa).
4. Baixa as imagens favoritas (miniaturas) e monta uma planilha .xlsx
   organizada, com abas: Observacoes (limpa), Taxa (resumo), Specimens
   (AntWeb) e Imagens (com thumbnails embutidos).
5. Exporta tudo para R: CSVs + script antweb/R/load_antweb.R que carrega os
   data.frames e (opcionalmente) salva um antweb_dataset.RData.

Uso
---
    # com rede (fala com o AntWeb real):
    python antweb/build.py --xlsx <planilha.xlsx>

    # offline / validação do pipeline (sem rede), usando fixtures mock:
    python antweb/build.py --xlsx <planilha.xlsx> --mock

    # limitar specimens por táxon (útil para amostra de demonstração):
    python antweb/build.py --xlsx <planilha.xlsx> --max-per-taxon 50 --download-images

Requisitos: pip install -r antweb/requirements.txt
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from antweb_client import AntWebClient, AntWebConfig, flatten_specimen  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
log = logging.getLogger("build")

OUT_DIR = HERE / "output"
IMG_DIR = HERE / "images"
R_DIR = HERE / "R"

# Colunas esperadas na planilha de observação focal.
OBS_COLUMNS = [
    "campanha", "area_verde", "cod_area", "ponto", "isca", "subfamily",
    "functional", "ant_genus", "ant_species", "comportamento", "distancia_m",
    "hora_inicio", "hora_comportamento", "direcao_remocao", "recrutamento",
    "dominancia", "abund", "obs",
]

NA_TOKENS = {"", "na", "n/a", "none", "null", "-", "nan"}


# --------------------------------------------------------------------- limpeza
def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _clean_num(v: Any) -> Optional[float]:
    s = _s(v).replace(",", ".")
    if s.lower() in NA_TOKENS:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _title_area(v: Any) -> str:
    s = _s(v)
    if s.lower() in NA_TOKENS:
        return ""
    # normaliza espaços e capitalização básica preservando acentos
    s = re.sub(r"\s+", " ", s).strip()
    return s[:1].upper() + s[1:] if s else s


def normalize_subfamily(v: Any) -> str:
    s = _s(v).lower()
    if s in NA_TOKENS:
        return ""
    return s.capitalize()


def normalize_genus(v: Any) -> str:
    s = _s(v)
    if s.lower() in NA_TOKENS:
        return ""
    return s.strip().capitalize()


def normalize_species(genus: str, v: Any) -> str:
    """Padroniza morfoespécie: 'Genero_spN' com underscore; espécies reais em minúsculo.

    Ex.: 'Pachycondyla Striata'  -> 'Pachycondyla_striata'
         'Crematogaster-sp4'      -> 'Crematogaster_sp4'
         'Camponotus_Senex'       -> 'Camponotus_senex'
    """
    s = _s(v)
    if s.lower() in NA_TOKENS:
        return ""
    # unifica separadores (espaço, hífen) em underscore
    s = re.sub(r"[\s\-]+", "_", s.strip())
    parts = s.split("_", 1)
    head = normalize_genus(parts[0]) if parts else ""
    tail = parts[1] if len(parts) > 1 else ""
    # morfoespécie 'spN' fica minúscula; epíteto específico real minúsculo
    tail = tail.lower()
    tail = re.sub(r"^sp0*(\d+)$", r"sp\1", tail)  # sp01 -> sp1
    if head and tail:
        return f"{head}_{tail}"
    return head or s


def is_morphospecies(species_norm: str) -> bool:
    return bool(re.search(r"_sp\d+$", species_norm or ""))


def species_epithet_for_antweb(species_norm: str) -> str:
    """Retorna o epíteto real para consulta na API, ou '' se for morfoespécie."""
    if not species_norm or is_morphospecies(species_norm):
        return ""
    parts = species_norm.split("_", 1)
    return parts[1].lower() if len(parts) > 1 else ""


def load_and_clean_observations(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [ _s(h).lower() for h in rows[0] ]

    def col(*names: str) -> int:
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    idx = {
        "campanha": col("campanha"),
        "area_verde": col("area_verde"),
        "cod_area": col("cod_are", "cod_area"),
        "ponto": col("ponto"),
        "isca": col("isca"),
        "subfamily": col("subfamily"),
        "functional": col("functional"),
        "ant_genus": col("ant_genus"),
        "ant_species": col("ant_species"),
        "comportamento": col("comportamento"),
        "distancia_m": col("distancia (m)", "distancia_m"),
        "hora_inicio": col("hora_inico", "hora_inicio"),
        "hora_comportamento": col("hora_comportamento"),
        "direcao_remocao": col("direcao_remocao"),
        "recrutamento": col("recrutamento"),
        "dominancia": col("dominancia"),
        "abund": col("abund"),
        "obs": col("obs:", "obs"),
    }

    cleaned: List[Dict[str, Any]] = []
    dropped = 0
    for r in rows[1:]:
        if not any(_s(c) for c in r):
            continue
        genus = normalize_genus(r[idx["ant_genus"]] if idx["ant_genus"] >= 0 else "")
        subf = normalize_subfamily(r[idx["subfamily"]] if idx["subfamily"] >= 0 else "")
        species = normalize_species(genus, r[idx["ant_species"]] if idx["ant_species"] >= 0 else "")
        # descarta linhas totalmente 'na' (sem gênero e sem subfamília úteis)
        if not genus and not subf:
            dropped += 1
            continue
        rec = {
            "campanha": _clean_num(r[idx["campanha"]]) if idx["campanha"] >= 0 else None,
            "area_verde": _title_area(r[idx["area_verde"]]) if idx["area_verde"] >= 0 else "",
            "cod_area": _s(r[idx["cod_area"]]).upper() if idx["cod_area"] >= 0 else "",
            "ponto": _clean_num(r[idx["ponto"]]) if idx["ponto"] >= 0 else None,
            "isca": _s(r[idx["isca"]]).upper() if idx["isca"] >= 0 else "",
            "subfamily": subf,
            "functional": _s(r[idx["functional"]]) if idx["functional"] >= 0 else "",
            "ant_genus": genus,
            "ant_species": species,
            "comportamento": _clean_num(r[idx["comportamento"]]) if idx["comportamento"] >= 0 else None,
            "distancia_m": _clean_num(r[idx["distancia_m"]]) if idx["distancia_m"] >= 0 else None,
            "hora_inicio": _s(r[idx["hora_inicio"]]) if idx["hora_inicio"] >= 0 else "",
            "hora_comportamento": _clean_num(r[idx["hora_comportamento"]]) if idx["hora_comportamento"] >= 0 else None,
            "direcao_remocao": _s(r[idx["direcao_remocao"]]).lower() if idx["direcao_remocao"] >= 0 else "",
            "recrutamento": _s(r[idx["recrutamento"]]).lower() if idx["recrutamento"] >= 0 else "",
            "dominancia": _s(r[idx["dominancia"]]).lower() if idx["dominancia"] >= 0 else "",
            "abund": _clean_num(r[idx["abund"]]) if idx["abund"] >= 0 else None,
            "obs": _s(r[idx["obs"]]) if idx["obs"] >= 0 else "",
            "is_morphospecies": is_morphospecies(species),
        }
        cleaned.append(rec)

    # dicionário meta_data (se existir)
    meta: List[Dict[str, str]] = []
    if "meta_data" in wb.sheetnames:
        for mr in wb["meta_data"].iter_rows(values_only=True):
            if _s(mr[0]) or (len(mr) > 1 and _s(mr[1])):
                meta.append({"campo": _s(mr[0]), "descricao": _s(mr[1]) if len(mr) > 1 else ""})

    log.info("Observações limpas: %d (descartadas %d linhas 'na')", len(cleaned), dropped)
    return cleaned, meta


def distinct_taxa(obs: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    seen: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for r in obs:
        g = r["ant_genus"]
        if not g:
            continue
        key = (r["subfamily"], g)
        entry = seen.setdefault(key, {
            "subfamily": r["subfamily"], "genus": g,
            "morphospecies": set(), "n_obs": 0,
        })
        entry["n_obs"] += 1
        if r["ant_species"]:
            entry["morphospecies"].add(r["ant_species"])
    out = []
    for (_sf, _g), e in sorted(seen.items()):
        out.append({
            "subfamily": e["subfamily"],
            "genus": e["genus"],
            "n_morphospecies": len(e["morphospecies"]),
            "morphospecies": ", ".join(sorted(e["morphospecies"])),
            "n_obs": e["n_obs"],
        })
    return out


# ------------------------------------------------------------------- AntWeb pull
def make_client(mock: bool, cfg: AntWebConfig) -> AntWebClient:
    if mock:
        from mock_session import MockSession
        return AntWebClient(cfg, session=MockSession(HERE / "mock_fixtures.json"))
    return AntWebClient(cfg)


def pull_specimens(client: AntWebClient, taxa: List[Dict[str, str]],
                   max_per_taxon: Optional[int]) -> List[Dict[str, Any]]:
    all_rows: List[Dict[str, Any]] = []
    for t in taxa:
        genus = t["genus"]
        if not genus:
            continue
        cfg = client.cfg
        old = cfg.max_records
        cfg.max_records = max_per_taxon
        log.info("AntWeb: puxando specimens de genus=%s ...", genus)
        try:
            for rec in client.iter_specimens(genus=genus.lower()):
                flat = flatten_specimen(rec)
                flat["study_subfamily"] = t["subfamily"]
                all_rows.append(flat)
        except Exception as exc:
            log.error("Falha ao puxar genus=%s: %s", genus, exc)
        finally:
            cfg.max_records = old
    log.info("Total de specimens AntWeb coletados: %d", len(all_rows))
    return all_rows


def download_favorite_images(specimens: List[Dict[str, Any]], limit: int,
                             mock: bool) -> Dict[str, str]:
    """Baixa até `limit` favorite images. Retorna {catalogNumber: caminho_local}."""
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    paths: Dict[str, str] = {}
    if mock:
        log.info("Modo mock: pulando download real de imagens.")
        return paths
    try:
        import requests
    except Exception:
        log.warning("requests indisponível; sem download de imagens.")
        return paths
    n = 0
    for s in specimens:
        url = s.get("favorite_image")
        cat = s.get("catalogNumber") or f"img{n}"
        if not url or n >= limit:
            continue
        dest = IMG_DIR / f"{cat}.jpg"
        try:
            resp = requests.get(url, timeout=60)
            resp.raise_for_status()
            dest.write_bytes(resp.content)
            paths[cat] = str(dest)
            n += 1
        except Exception as exc:
            log.warning("Falha ao baixar imagem %s: %s", url, exc)
    log.info("Imagens baixadas: %d", len(paths))
    return paths


# ------------------------------------------------------------------- planilha
def build_xlsx(obs, meta, taxa, specimens, image_paths, out_path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F5C3F")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def autosize(ws, maxw=48):
        for col in ws.columns:
            length = max((len(_s(c.value)) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxw, max(10, length + 2))

    # aba 1: Observações limpas
    ws1 = wb.active
    ws1.title = "Observacoes"
    cols1 = OBS_COLUMNS + ["is_morphospecies"]
    ws1.append(cols1)
    for r in obs:
        ws1.append([r.get(c if c != "is_morphospecies" else "is_morphospecies") for c in cols1])
    style_header(ws1, len(cols1))
    autosize(ws1)

    # aba 2: Taxa (resumo)
    ws2 = wb.create_sheet("Taxa")
    cols2 = ["subfamily", "genus", "n_morphospecies", "morphospecies", "n_obs"]
    ws2.append(cols2)
    for t in taxa:
        ws2.append([t.get(c) for c in cols2])
    style_header(ws2, len(cols2))
    autosize(ws2)

    # aba 3: Specimens (AntWeb)
    ws3 = wb.create_sheet("Specimens")
    cols3 = ["catalogNumber", "scientific_name", "family", "study_subfamily",
             "subfamily", "genus", "species", "caste", "typeStatus", "country",
             "stateProvince", "locality", "habitat", "minimumElevationInMeters",
             "decimal_latitude", "decimal_longitude", "n_images",
             "favorite_image", "url", "image_urls"]
    ws3.append(cols3)
    for s in specimens:
        row = [s.get(c) for c in cols3]
        ws3.append(row)
        # hyperlink na coluna url
        cell = ws3.cell(row=ws3.max_row, column=cols3.index("url") + 1)
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = Font(color="0563C1", underline="single")
    style_header(ws3, len(cols3))
    autosize(ws3)

    # aba 4: Imagens (thumbnails embutidos das favorite images)
    ws4 = wb.create_sheet("Imagens")
    ws4.append(["catalogNumber", "scientific_name", "favorite_image", "thumbnail"])
    style_header(ws4, 4)
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 28
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 24
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        XLImage = None
    rownum = 2
    for s in specimens:
        cat = s.get("catalogNumber")
        ws4.cell(row=rownum, column=1, value=cat)
        ws4.cell(row=rownum, column=2, value=s.get("scientific_name"))
        c3 = ws4.cell(row=rownum, column=3, value=s.get("favorite_image"))
        if c3.value:
            c3.hyperlink = c3.value
            c3.font = Font(color="0563C1", underline="single")
        local = image_paths.get(cat)
        if local and XLImage and os.path.exists(local):
            try:
                img = XLImage(local)
                img.width, img.height = 120, 90
                ws4.row_dimensions[rownum].height = 70
                ws4.add_image(img, f"D{rownum}")
            except Exception as exc:
                log.warning("Não embutiu imagem %s: %s", local, exc)
        rownum += 1

    # aba 5: dicionário meta_data
    if meta:
        ws5 = wb.create_sheet("Dicionario")
        ws5.append(["campo", "descricao"])
        for m in meta:
            ws5.append([m["campo"], m["descricao"]])
        style_header(ws5, 2)
        autosize(ws5, maxw=80)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Planilha organizada salva em %s", out_path)


# ------------------------------------------------------------------- exports R
def write_csvs(obs, taxa, specimens) -> Dict[str, Path]:
    import csv
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    files = {}

    def dump(name, rows, cols):
        p = OUT_DIR / name
        with p.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow(r)
        files[name] = p
        return p

    dump("observacoes.csv", obs, OBS_COLUMNS + ["is_morphospecies"])
    dump("taxa.csv", taxa, ["subfamily", "genus", "n_morphospecies", "morphospecies", "n_obs"])
    dump("specimens.csv", specimens,
         ["catalogNumber", "scientific_name", "family", "study_subfamily", "subfamily",
          "genus", "species", "caste", "typeStatus", "country", "stateProvince",
          "locality", "habitat", "minimumElevationInMeters", "decimal_latitude",
          "decimal_longitude", "n_images", "favorite_image", "url", "image_urls"])
    log.info("CSVs escritos em %s", OUT_DIR)
    return files


def write_r_script():
    R_DIR.mkdir(parents=True, exist_ok=True)
    script = r'''# load_antweb.R
# Carrega os dados organizados (Serrapilheira + AntWeb) em data.frames de R.
# Gerado por antweb/build.py. Rode a partir da raiz do repositório:
#     source("antweb/R/load_antweb.R")
#
# Requer apenas base R (utils::read.csv). Opcionalmente usa jsonlite/httr
# para (re)consultar a API do AntWeb ao vivo — veranexo no fim do arquivo.

antweb_dir <- file.path("antweb", "output")
if (!dir.exists(antweb_dir)) {
  # fallback: rodando de dentro de antweb/R
  antweb_dir <- normalizePath(file.path("..", "output"), mustWork = FALSE)
}

read_or_empty <- function(name) {
  path <- file.path(antweb_dir, name)
  if (file.exists(path)) {
    utils::read.csv(path, stringsAsFactors = FALSE, encoding = "UTF-8")
  } else {
    warning(sprintf("Arquivo nao encontrado: %s", path))
    data.frame()
  }
}

observacoes <- read_or_empty("observacoes.csv")
taxa        <- read_or_empty("taxa.csv")
specimens   <- read_or_empty("specimens.csv")

# Tipagem util
num_cols <- c("campanha", "ponto", "comportamento", "distancia_m",
              "hora_comportamento", "abund")
for (nm in intersect(num_cols, names(observacoes))) {
  observacoes[[nm]] <- suppressWarnings(as.numeric(observacoes[[nm]]))
}
for (nm in intersect(c("subfamily", "ant_genus", "ant_species", "area_verde",
                       "isca", "functional"), names(observacoes))) {
  observacoes[[nm]] <- as.factor(observacoes[[nm]])
}

cat(sprintf("Observacoes: %d linhas\n", nrow(observacoes)))
cat(sprintf("Taxa: %d\n", nrow(taxa)))
cat(sprintf("Specimens (AntWeb): %d\n", nrow(specimens)))

# Salva um .RData consolidado para reuso rapido
save(observacoes, taxa, specimens,
     file = file.path(antweb_dir, "antweb_dataset.RData"))
cat(sprintf("Salvo: %s\n", file.path(antweb_dir, "antweb_dataset.RData")))

# ------------------------------------------------------------------------------
# ANEXO (opcional): consultar a API do AntWeb ao vivo direto do R.
# Descomente e instale: install.packages(c("httr", "jsonlite"))
#
# antweb_fetch <- function(genus, limit = 100, offset = 0) {
#   resp <- httr::GET("https://www.antweb.org/api/v2/",
#                     query = list(genus = tolower(genus),
#                                  limit = limit, offset = offset))
#   httr::stop_for_status(resp)
#   jsonlite::fromJSON(httr::content(resp, as = "text", encoding = "UTF-8"),
#                      simplifyVector = FALSE)
# }
# pheidole <- antweb_fetch("pheidole", limit = 10)
# str(pheidole$count)
'''
    p = R_DIR / "load_antweb.R"
    p.write_text(script, encoding="utf-8")
    log.info("Script R salvo em %s", p)


# ------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Pipeline AntWeb + Serrapilheira")
    ap.add_argument("--xlsx", required=True, help="Planilha de observação focal (.xlsx)")
    ap.add_argument("--mock", action="store_true", help="Usa fixtures locais (sem rede)")
    ap.add_argument("--max-per-taxon", type=int, default=None,
                    help="Limite de specimens por gênero (None=todos)")
    ap.add_argument("--download-images", action="store_true",
                    help="Baixa as favorite images (miniaturas)")
    ap.add_argument("--image-limit", type=int, default=40,
                    help="Máx. de favorite images a baixar/embutir")
    ap.add_argument("--out", default=str(OUT_DIR / "Serrapilheira_AntWeb_ORGANIZADA.xlsx"))
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        log.error("Planilha não encontrada: %s", xlsx)
        sys.exit(1)

    obs, meta = load_and_clean_observations(xlsx)
    taxa = distinct_taxa(obs)
    log.info("Táxons (gêneros) no estudo: %s", ", ".join(t["genus"] for t in taxa))

    cfg = AntWebConfig(max_records=args.max_per_taxon)
    client = make_client(args.mock, cfg)
    specimens = pull_specimens(client, taxa, args.max_per_taxon)

    image_paths = {}
    if args.download_images:
        image_paths = download_favorite_images(specimens, args.image_limit, args.mock)

    build_xlsx(obs, meta, taxa, specimens, image_paths, Path(args.out))
    write_csvs(obs, taxa, specimens)
    write_r_script()

    # manifesto JSON (útil para o app / integração)
    manifest = {
        "n_observacoes": len(obs),
        "n_taxa": len(taxa),
        "n_specimens": len(specimens),
        "generos": [t["genus"] for t in taxa],
        "outputs": {
            "xlsx": args.out,
            "csv": [str(OUT_DIR / n) for n in ("observacoes.csv", "taxa.csv", "specimens.csv")],
            "r_script": str(R_DIR / "load_antweb.R"),
        },
        "mock": args.mock,
    }
    (OUT_DIR / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2))
    log.info("Concluído. Manifesto: %s", OUT_DIR / "manifest.json")


if __name__ == "__main__":
    main()
